import os
import logging

from openpyxl import load_workbook


class Utils:
    def __init__(self):
        self.logger = self.custom_logger()

    def custom_logger(module_name, logLevel=logging.DEBUG):
        logger = logging.getLogger(module_name)
        if not logger.handlers:
            logger.setLevel(logLevel)
            log_dir = f"../Logs/{module_name}_logs"
            os.makedirs(log_dir, exist_ok=True)
            log_file = os.path.join(log_dir, f"{module_name}.log")
            fh = logging.FileHandler(log_file)
            formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(name)s - %(message)s')
            fh.setFormatter(formatter)
            logger.addHandler(fh)
        return logger

    def read_xlsx(file_name, sheet):
        datalist = []
        try:
            wb = load_workbook(filename=file_name)
            sh = wb[sheet]
            row_ct = sh.max_row
            col_ct = sh.max_column
            # Starting the range from 2 because first row content the headers
            for i in range(2, row_ct + 1):
                row = []
                for j in range(1, col_ct + 1):
                    row.append(sh.cell(row=i, column=j).value)
                datalist.append(row)
            return datalist
        except FileNotFoundError:
            raise Exception(f"File not located: {file_name}")



